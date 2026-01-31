"""XLSForm survey importer."""

from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from voicetest.importers.base import ImporterInfo
from voicetest.models.agent import (
    AgentGraph,
    AgentNode,
    Transition,
    TransitionCondition,
)


class XLSFormImporter:
    """Import XLSForm Excel files as voice agent flows.

    XLSForm is a standard for authoring forms in Excel. This importer
    converts survey questions into conversation nodes where:
    - Each question becomes a node that asks the user for input
    - Skip logic (relevant column) becomes transition conditions
    - Select questions branch based on user choices
    """

    @property
    def source_type(self) -> str:
        return "xlsform"

    def get_info(self) -> ImporterInfo:
        return ImporterInfo(
            source_type="xlsform",
            description="Import XLSForm Excel surveys as voice agent flows",
            file_patterns=["*.xlsx", "*.xls"],
        )

    def can_import(self, path_or_config: str | Path | dict) -> bool:
        """Detect XLSForm by checking for survey sheet with required columns."""
        if isinstance(path_or_config, dict):
            return False

        path = Path(path_or_config)
        if not path.exists() or path.suffix.lower() not in (".xlsx", ".xls"):
            return False

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if "survey" not in wb.sheetnames:
                return False

            survey_sheet = wb["survey"]
            headers = [cell.value for cell in next(survey_sheet.iter_rows(max_row=1))]
            required_cols = {"type", "name", "label"}
            return required_cols.issubset(set(h.lower() if h else "" for h in headers))
        except Exception:
            return False

    def import_agent(self, path_or_config: str | Path | dict) -> AgentGraph:
        """Convert XLSForm to AgentGraph."""
        if isinstance(path_or_config, dict):
            raise ValueError("XLSForm importer requires a file path, not a dict")

        path = Path(path_or_config)
        wb = load_workbook(path, data_only=True)

        survey_rows = self._parse_sheet(wb["survey"])
        choices = self._parse_choices(wb) if "choices" in wb.sheetnames else {}
        settings = self._parse_settings(wb) if "settings" in wb.sheetnames else {}

        return self._build_graph(survey_rows, choices, settings, path.stem)

    def _parse_sheet(self, sheet) -> list[dict[str, Any]]:
        """Parse a worksheet into list of row dicts."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        return [
            {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
            for row in rows[1:]
            if any(cell for cell in row)
        ]

    def _parse_choices(self, wb) -> dict[str, list[dict[str, str]]]:
        """Parse choices sheet into {list_name: [{name, label}, ...]}."""
        choices_rows = self._parse_sheet(wb["choices"])
        choices: dict[str, list[dict[str, str]]] = {}

        for row in choices_rows:
            list_name = row.get("list_name") or row.get("list name", "")
            name = row.get("name", "")
            label = row.get("label", "")

            if list_name and name:
                if list_name not in choices:
                    choices[list_name] = []
                choices[list_name].append({"name": str(name), "label": str(label or name)})

        return choices

    def _parse_settings(self, wb) -> dict[str, Any]:
        """Parse settings sheet."""
        settings_rows = self._parse_sheet(wb["settings"])
        return settings_rows[0] if settings_rows else {}

    def _build_graph(
        self,
        survey_rows: list[dict[str, Any]],
        choices: dict[str, list[dict[str, str]]],
        settings: dict[str, Any],
        form_name: str,
    ) -> AgentGraph:
        """Build AgentGraph from parsed XLSForm data."""
        nodes: dict[str, AgentNode] = {}
        question_order: list[str] = []
        relevance_map: dict[str, str] = {}

        # First pass: create nodes for each question
        group_stack: list[str] = []
        for row in survey_rows:
            q_type = str(row.get("type", "")).strip().lower()
            q_name = str(row.get("name", "")).strip()
            q_label = str(row.get("label", "")).strip()
            q_relevant = row.get("relevant", "")
            q_hint = row.get("hint", "")
            q_required = row.get("required", "")

            if not q_name:
                continue

            # Handle group start/end
            if q_type == "begin_group" or q_type == "begin group":
                group_stack.append(q_name)
                continue
            elif q_type == "end_group" or q_type == "end group":
                if group_stack:
                    group_stack.pop()
                continue

            # Skip non-question types
            if q_type in ("calculate", "hidden", "start", "end", "deviceid", "today"):
                continue

            # Build state prompt for this node
            state_prompt = self._build_node_instructions(
                q_type, q_name, q_label, q_hint, q_required, choices
            )

            if not state_prompt:
                continue

            nodes[q_name] = AgentNode(
                id=q_name,
                state_prompt=state_prompt,
                tools=[],
                transitions=[],
                metadata={
                    "xlsform_type": q_type,
                    "group": "/".join(group_stack) if group_stack else None,
                },
            )
            question_order.append(q_name)

            if q_relevant:
                relevance_map[q_name] = str(q_relevant)

        # Second pass: create transitions
        for i, q_name in enumerate(question_order):
            node = nodes[q_name]
            next_questions = self._find_next_questions(q_name, i, question_order, relevance_map)

            for next_q, condition in next_questions:
                if next_q in nodes:
                    node.transitions.append(
                        Transition(
                            target_node_id=next_q,
                            condition=TransitionCondition(
                                type="llm_prompt" if condition else "always",
                                value=condition or "User has answered the question",
                            ),
                            description=condition or f"Continue to {next_q}",
                        )
                    )

        # Add end transition to last node
        if question_order and question_order[-1] in nodes:
            last_node = nodes[question_order[-1]]
            if not last_node.transitions:
                last_node.transitions.append(
                    Transition(
                        target_node_id="__end__",
                        condition=TransitionCondition(
                            type="always",
                            value="Survey complete",
                        ),
                        description="End of survey",
                    )
                )

        # Add end node
        nodes["__end__"] = AgentNode(
            id="__end__",
            state_prompt=(
                "Thank the user for completing the survey. "
                "Summarize their responses if appropriate and end the conversation politely."
            ),
            tools=[],
            transitions=[],
            metadata={"xlsform_type": "end"},
        )

        entry_node = question_order[0] if question_order else "__end__"

        return AgentGraph(
            nodes=nodes,
            entry_node_id=entry_node,
            source_type="xlsform",
            source_metadata={
                "form_name": settings.get("form_title", form_name),
                "form_id": settings.get("form_id", form_name),
                "general_prompt": "",  # XLSForm surveys don't have separate general prompt
            },
        )

    def _build_node_instructions(
        self,
        q_type: str,
        q_name: str,
        q_label: str,
        q_hint: str,
        q_required: str,
        choices: dict[str, list[dict[str, str]]],
    ) -> str:
        """Build agent instructions for a question node."""
        if not q_label:
            return ""

        parts = [f'Ask the user: "{q_label}"']

        if q_hint:
            parts.append(f"If they need help, explain: {q_hint}")

        # Add type-specific instructions
        if q_type == "integer":
            parts.append("Expect a whole number response.")
        elif q_type == "decimal":
            parts.append("Expect a numeric response (can include decimals).")
        elif q_type == "date":
            parts.append("Expect a date response.")
        elif q_type == "time":
            parts.append("Expect a time response.")
        elif q_type == "datetime":
            parts.append("Expect a date and time response.")
        elif q_type.startswith("select_one ") or q_type.startswith("select one "):
            list_name = q_type.split(" ", 1)[1].strip()
            if list_name in choices:
                options = choices[list_name]
                option_str = ", ".join(f"'{opt['label']}'" for opt in options)
                parts.append(f"The valid options are: {option_str}.")
                parts.append("Accept only one of these options.")
        elif q_type.startswith("select_multiple ") or q_type.startswith("select multiple "):
            list_name = q_type.split(" ", 1)[1].strip()
            if list_name in choices:
                options = choices[list_name]
                option_str = ", ".join(f"'{opt['label']}'" for opt in options)
                parts.append(f"The valid options are: {option_str}.")
                parts.append("The user can select multiple options.")
        elif q_type == "note":
            return f'Tell the user: "{q_label}". Then continue to the next question.'

        if q_required and str(q_required).lower() in ("yes", "true", "1"):
            parts.append("This question is required - do not continue until answered.")

        parts.append(f"Store their response as '{q_name}'.")

        return "\n".join(parts)

    def _find_next_questions(
        self,
        current_name: str,
        current_index: int,
        question_order: list[str],
        relevance_map: dict[str, str],
    ) -> list[tuple[str, str]]:
        """Find possible next questions with their conditions.

        Returns list of (next_question_name, condition_description) tuples.
        """
        next_questions: list[tuple[str, str]] = []

        # Look at all following questions
        for i in range(current_index + 1, len(question_order)):
            next_q = question_order[i]
            relevant = relevance_map.get(next_q, "")

            if relevant:
                condition = self._parse_relevant_condition(relevant, current_name)
                if condition:
                    next_questions.append((next_q, condition))
            else:
                # No relevance condition - this is the default next
                next_questions.append((next_q, ""))
                break

        return next_questions if next_questions else []

    def _parse_relevant_condition(self, relevant: str, current_var: str) -> str:
        """Convert XLSForm relevant expression to natural language condition.

        Examples:
            ${has_appt} = 'yes' -> "User answered 'yes' to has_appt"
            selected(${symptoms}, 'fever') -> "User selected 'fever' for symptoms"
            ${age} > 18 -> "age is greater than 18"
        """
        # Extract variable references
        var_pattern = r"\$\{([^}]+)\}"
        vars_used = re.findall(var_pattern, relevant)

        # Simple equality: ${var} = 'value'
        eq_match = re.match(r"\$\{([^}]+)\}\s*=\s*['\"]([^'\"]+)['\"]", relevant)
        if eq_match:
            var_name, value = eq_match.groups()
            return f"User answered '{value}' to {var_name}"

        # Not equal: ${var} != 'value'
        neq_match = re.match(r"\$\{([^}]+)\}\s*!=\s*['\"]([^'\"]+)['\"]", relevant)
        if neq_match:
            var_name, value = neq_match.groups()
            return f"User did not answer '{value}' to {var_name}"

        # Selected function: selected(${var}, 'value')
        sel_pattern = r"selected\s*\(\s*\$\{([^}]+)\}\s*,\s*['\"]([^'\"]+)['\"]\s*\)"
        sel_match = re.match(sel_pattern, relevant)
        if sel_match:
            var_name, value = sel_match.groups()
            return f"User selected '{value}' for {var_name}"

        # Numeric comparisons
        num_match = re.match(r"\$\{([^}]+)\}\s*([<>=!]+)\s*(\d+)", relevant)
        if num_match:
            var_name, op, value = num_match.groups()
            op_text = {
                ">": "is greater than",
                "<": "is less than",
                ">=": "is at least",
                "<=": "is at most",
                "=": "equals",
                "!=": "does not equal",
            }.get(op, op)
            return f"{var_name} {op_text} {value}"

        # Fallback: return a generic condition
        if vars_used:
            return f"Based on {', '.join(vars_used)}: {relevant}"

        return relevant
