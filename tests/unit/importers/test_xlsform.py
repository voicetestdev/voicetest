"""Tests for XLSForm importer."""

from pathlib import Path

from openpyxl import Workbook
import pytest

from voicetest.importers.xlsform import XLSFormImporter


@pytest.fixture
def xlsform_fixtures_dir() -> Path:
    """Return path to XLSForm test fixtures."""
    return Path(__file__).parent.parent / "fixtures" / "xlsform"


@pytest.fixture
def sample_xlsform(xlsform_fixtures_dir: Path) -> Path:
    """Return path to sample XLSForm."""
    return xlsform_fixtures_dir / "sample_survey.xlsx"


@pytest.fixture
def importer() -> XLSFormImporter:
    """Create an XLSForm importer."""
    return XLSFormImporter()


@pytest.fixture
def minimal_xlsform(tmp_path: Path) -> Path:
    """Create a minimal XLSForm for testing."""
    wb = Workbook()
    survey = wb.active
    survey.title = "survey"
    survey.append(["type", "name", "label"])
    survey.append(["text", "q1", "What is your name?"])
    survey.append(["text", "q2", "What is your favorite color?"])

    path = tmp_path / "minimal.xlsx"
    wb.save(path)
    return path


class TestXLSFormImporterDetection:
    """Tests for XLSForm format detection."""

    def test_source_type(self, importer):
        assert importer.source_type == "xlsform"

    def test_get_info(self, importer):
        info = importer.get_info()
        assert info.source_type == "xlsform"
        assert "*.xlsx" in info.file_patterns

    def test_can_import_valid_xlsform(self, importer, sample_xlsform):
        assert importer.can_import(sample_xlsform)

    def test_can_import_minimal_xlsform(self, importer, minimal_xlsform):
        assert importer.can_import(minimal_xlsform)

    def test_cannot_import_dict(self, importer):
        assert not importer.can_import({"type": "text", "name": "q1"})

    def test_cannot_import_nonexistent_file(self, importer):
        assert not importer.can_import("/nonexistent/file.xlsx")

    def test_cannot_import_non_excel_file(self, importer, tmp_path):
        json_file = tmp_path / "test.json"
        json_file.write_text('{"test": true}')
        assert not importer.can_import(json_file)

    def test_cannot_import_excel_without_survey_sheet(self, importer, tmp_path):
        wb = Workbook()
        wb.active.title = "data"
        path = tmp_path / "no_survey.xlsx"
        wb.save(path)
        assert not importer.can_import(path)

    def test_cannot_import_excel_without_required_columns(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["question", "answer"])
        path = tmp_path / "wrong_columns.xlsx"
        wb.save(path)
        assert not importer.can_import(path)


class TestXLSFormImporterBasic:
    """Tests for basic XLSForm importing."""

    def test_import_creates_graph(self, importer, minimal_xlsform):
        graph = importer.import_agent(minimal_xlsform)
        assert graph is not None
        assert graph.source_type == "xlsform"

    def test_import_creates_nodes_for_questions(self, importer, minimal_xlsform):
        graph = importer.import_agent(minimal_xlsform)
        assert "q1" in graph.nodes
        assert "q2" in graph.nodes

    def test_import_sets_entry_node(self, importer, minimal_xlsform):
        graph = importer.import_agent(minimal_xlsform)
        assert graph.entry_node_id == "q1"

    def test_import_creates_end_node(self, importer, minimal_xlsform):
        graph = importer.import_agent(minimal_xlsform)
        assert "__end__" in graph.nodes

    def test_node_instructions_contain_label(self, importer, minimal_xlsform):
        graph = importer.import_agent(minimal_xlsform)
        assert "What is your name?" in graph.nodes["q1"].instructions

    def test_nodes_have_transitions(self, importer, minimal_xlsform):
        graph = importer.import_agent(minimal_xlsform)
        assert len(graph.nodes["q1"].transitions) > 0


class TestXLSFormImporterSample:
    """Tests using the sample XLSForm fixture."""

    def test_import_sample_survey(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        assert graph is not None

    def test_sample_has_expected_nodes(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        expected_nodes = ["name", "age", "has_appointment", "appointment_date", "visit_reason"]
        for node_name in expected_nodes:
            assert node_name in graph.nodes, f"Missing node: {node_name}"

    def test_sample_preserves_form_metadata(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        assert graph.source_metadata.get("form_name") == "Appointment Scheduler"
        assert graph.source_metadata.get("form_id") == "appointment_form"

    def test_select_one_includes_options(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        instructions = graph.nodes["has_appointment"].instructions
        assert "Yes" in instructions
        assert "No" in instructions

    def test_conditional_question_has_condition_in_transition(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        has_appt_node = graph.nodes["has_appointment"]

        transition_targets = [t.target_node_id for t in has_appt_node.transitions]
        assert "appointment_date" in transition_targets or "visit_reason" in transition_targets

    def test_required_field_noted_in_instructions(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        assert "required" in graph.nodes["name"].instructions.lower()

    def test_integer_type_noted_in_instructions(self, importer, sample_xlsform):
        graph = importer.import_agent(sample_xlsform)
        assert "number" in graph.nodes["age"].instructions.lower()


class TestXLSFormImporterChoices:
    """Tests for choice list handling."""

    def test_select_one_lists_all_options(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label"])
        survey.append(["select_one colors", "fav_color", "What is your favorite color?"])

        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label"])
        choices.append(["colors", "red", "Red"])
        choices.append(["colors", "blue", "Blue"])
        choices.append(["colors", "green", "Green"])

        path = tmp_path / "colors.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)
        instructions = graph.nodes["fav_color"].instructions

        assert "Red" in instructions
        assert "Blue" in instructions
        assert "Green" in instructions

    def test_select_multiple_allows_multiple(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label"])
        survey.append(["select_multiple symptoms", "symptoms", "What symptoms do you have?"])

        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label"])
        choices.append(["symptoms", "fever", "Fever"])
        choices.append(["symptoms", "cough", "Cough"])

        path = tmp_path / "symptoms.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)
        instructions = graph.nodes["symptoms"].instructions

        assert "multiple" in instructions.lower()


class TestXLSFormImporterRelevance:
    """Tests for skip logic / relevance conditions."""

    def test_relevance_creates_conditional_transition(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label", "relevant"])
        survey.append(["select_one yn", "has_pet", "Do you have a pet?", ""])
        survey.append(["text", "pet_name", "What is your pet's name?", "${has_pet} = 'yes'"])

        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label"])
        choices.append(["yn", "yes", "Yes"])
        choices.append(["yn", "no", "No"])

        path = tmp_path / "pets.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)
        has_pet_transitions = graph.nodes["has_pet"].transitions

        conditions = [t.condition.value for t in has_pet_transitions]
        assert any("yes" in c.lower() for c in conditions)

    def test_numeric_relevance_parsed(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label", "relevant"])
        survey.append(["integer", "age", "How old are you?", ""])
        survey.append(["text", "adult_q", "Adult-only question", "${age} >= 18"])

        path = tmp_path / "age.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)
        age_transitions = graph.nodes["age"].transitions

        conditions = [t.condition.value for t in age_transitions]
        assert any("18" in c for c in conditions)


class TestXLSFormImporterGroups:
    """Tests for group handling."""

    def test_groups_tracked_in_metadata(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label"])
        survey.append(["begin_group", "personal_info", "Personal Information"])
        survey.append(["text", "name", "What is your name?"])
        survey.append(["integer", "age", "How old are you?"])
        survey.append(["end_group", "", ""])

        path = tmp_path / "groups.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)

        assert graph.nodes["name"].metadata.get("group") == "personal_info"
        assert graph.nodes["age"].metadata.get("group") == "personal_info"


class TestXLSFormImporterEdgeCases:
    """Tests for edge cases."""

    def test_empty_survey_creates_minimal_graph(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label"])

        path = tmp_path / "empty.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)
        assert graph is not None
        assert "__end__" in graph.nodes

    def test_note_type_creates_informational_node(self, importer, tmp_path):
        wb = Workbook()
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label"])
        survey.append(["note", "welcome", "Welcome to this survey!"])

        path = tmp_path / "note.xlsx"
        wb.save(path)

        graph = importer.import_agent(path)
        assert "welcome" in graph.nodes
        assert "Welcome" in graph.nodes["welcome"].instructions

    def test_import_raises_for_dict_input(self, importer):
        with pytest.raises(ValueError, match="requires a file path"):
            importer.import_agent({"type": "text"})
